from plistlib import InvalidFileException
import openpyxl
from datetime import datetime
import os


class sheetValidation:

    """
    A class for performing validation
    of excel file and returns result in console and file in txt and xsxl.

    Attributes:
        path (str): File path for validation.
        error_len (list): List of errors for the length cell value.
        error_default (list): List of errors for the default column values.
        error_date (list):   List of errors for the data column.
        default_coordinate: List of cell coordinates of default errors.
        len_coordinate:  List of cell coordinates of lenght errors.
        date_coordinate: List of cell coordinates of date errors.
    """

    def __init__(self, path):
        self.path = path
        self.error_len = []
        self.error_default = []
        self.error_date = []
        self.default_coordinate = []
        self.len_coordinate = []
        self.date_coordinate = []

        # Check if the file path is valid and load the workbook
        self.check_sheet_path()

        # Set up variables for writing output file
        self.folder_path = os.path.expanduser("~/Downloads/")
        self.file_name = path.split("/")[-1]
        self.split_file_name = os.path.splitext(self.file_name)
        self.date_time_str = datetime.now().strftime("%Y-%m-%d-%H-%M-%S-%f")

    def check_sheet_path(self):
        """
        Check if the file path is valid and load the workbook

        Args:
        self : obj -- Instance of the class containing the file path and workbook.

        Returns:
        str: Return a message indicating if the file was processed successfully or not.

        Raises:
        InvalidFileException --  throws invalid format of file path
        Exception -- throws error working with file
        """
        try:
            self.workbook = openpyxl.load_workbook(self.path)
        except InvalidFileException:
            return "Invalid file format!"
        except Exception as e:
            return f"An error occurred: {e}"
        else:
            self.worksheet = self.workbook.active
        finally:
            return "Done processing the file."

    def validate_len_of_first_column(self):
        """
        Check if cell values in the first column have a length of 3

        Args:
        self (obj): Instance of the class containing the file path and workbook.

        Returns:
        list: List of invalid cell messages or an empty list if all cells are valid.

        """

        for row in self.worksheet.iter_rows(min_row=2):
            cell_value = row[0].value

            if isinstance(cell_value, (int, float)):
                cell_value_str = str(int(cell_value))
            else:
                cell_value_str = str(cell_value)

            if cell_value is not None and len(cell_value_str) != 3:
                self.len_coordinate.append(row[0].coordinate)
                self.error_len.append(
                    f"(*) Cell {row[0].coordinate} has an invalid length"
                )
        return self.error_len

    def validate_default_column(self):
        """
        Check if cell values in the Default column have a value Y or N

        Args:
        self : obj -- Instance of the class containing the file path and workbook.

        Returns:
        list: List of invalid cell messages or an empty list if all cells are valid.
        """
        default_index = None
        for row in self.worksheet.iter_rows(min_row=1, max_row=1):
            default_index = None
            for cell in row:
                if cell.value == "Default":
                    default_index = cell.column
                    break
        if default_index is not None:
            max_row = self.worksheet.max_row
            for row in self.worksheet.iter_rows(min_row=2, max_row=max_row):
                for cell in row:
                    if (
                        cell.column == default_index
                        and cell.value is not None
                        and cell.value not in ["Y", "N"]
                    ):
                        self.default_coordinate.append(cell.coordinate)
                        self.error_default.append(
                            f"(*) Cell {cell.coordinate} has an invalid value"
                        )

        return self.error_default

    def validate_date(self):
        """
        Check if cell date is valid in this century, between 2000 and current date.
        If date do not corenspond with this interval and is not current year, returns error message.

        Args:
        self : obj -- Instance of the class containing the file path and workbook.

        Returns:
        list: List of invalid cell messages or an empty list if all cells are valid.
        """

        for row in self.worksheet.iter_rows():
            current_year = datetime.now().year
            for cell in row:
                if cell.value is not None and isinstance(cell.value, datetime) == True:
                    if cell.value.year < 2000 or cell.value.year > current_year:
                        self.date_coordinate.append(cell.coordinate)
                        self.error_date.append(
                            f"(*) Cell {cell.coordinate} has an invalid date"
                        )

        return self.error_date

    def write_to_file(self):
        """
        Writes all errors in file with txt exstension.

        Args:
            self : obj -- Instance of the class containing the file path and workbook.

        Returns:
           file: Returns file with corespondent cell coordinate and error.

        """
        generete_output_file_name = f"{self.split_file_name[0]}{self.date_time_str}.txt"
        full_path = os.path.join(self.folder_path, generete_output_file_name)
        with open(full_path, "w") as f:
            f.write("\n".join(self.error_date + self.error_len + self.error_default))
        os.startfile(full_path)

    def write_to_excel(self):
        """
        Writes all errors in selected excel file adding column
        at end with errors. In rows are errors with cell coordinate for that error.

        Args:
        self : obj -- Instance of the class containing the file path and workbook.

        Returns:
        file: Returns selected excel file with column and corespondent
                cell coordinate with error.

        """
        if "Errors" not in [cell.value for cell in self.worksheet[1]]:
            new_col_1 = self.worksheet.max_column + 1
            self.worksheet.cell(row=1, column=new_col_1, value="Errors")
        else:
            new_col_1 = [cell.value for cell in self.worksheet[1]].index("Errors") + 1

        errors = []
        for row in self.worksheet.iter_rows(min_row=2):
            error_row = []
            for cell in row:
                if cell.value is not None:
                    if str(cell.coordinate) in self.date_coordinate:
                        error_row.append(f"{cell.coordinate} wrong date")
                    elif str(cell.coordinate) in self.len_coordinate:
                        error_row.append(f"{cell.coordinate} wrong len")
                    elif str(cell.coordinate) in self.default_coordinate:
                        error_row.append(f"{cell.coordinate} wrong default")
                    else:
                        error_row.append(f"{cell.coordinate} Correct")
            if error_row:
                errors.append(error_row)
        for row_idx, error_row in enumerate(errors, start=2):
            error_string = "\n".join(error_row)
            self.worksheet.cell(row=row_idx, column=new_col_1, value=error_string)

        generete_output_file_name = (
            f"{self.split_file_name[0]}{self.date_time_str}{self.split_file_name[1]}"
        )
        full_path = os.path.join(self.folder_path, generete_output_file_name)
        self.workbook.save(full_path)
        os.startfile(full_path)

    def run(self):
        """
        Runs all validation methods on the workbook.

        Args:
        self : obj -- Instance of the class containing the file path and workbook.

        """
        self.validate_len_of_first_column()
        self.validate_default_column()
        self.validate_date()
